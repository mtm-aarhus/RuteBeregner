"""
Script til at oprette standardiseret Excel skabelon for Jord Transport systemet.
Dette script genererer en Excel fil med beskyttede headers og data validering.
"""
import os
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows

def create_excel_template():
    """
    Opretter en standardiseret Excel skabelon med:
    1. Obligatoriske felter for start/slut adresser
    2. Valgfrie felter for transport detaljer  
    3. Beskyttede headers
    4. Data validering
    5. Instruktionsark
    """
    
    # Opret ny workbook
    wb = Workbook()
    
    # Fjern default worksheet
    wb.remove(wb.active)
    
    # 1. Opret Data ark
    data_sheet = wb.create_sheet("Data", 0)
    
    # Definer kolonner
    mandatory_columns = [
        "Adresse",           # Start adresse gade/vej
        "Postnummer",        # Start adresse postnummer  
        "PostDistrikt",      # Start adresse by
        "SlutAdresse"        # Fleksibel slutadresse (adresse, koordinater eller anlæg ID)
    ]
    
    optional_columns = [
        "Navn",              # Virksomheds/projekt navn
        "Dato",              # Transport dato
        "KøretøjsType",      # Type køretøj
        "LastVægt",          # Vægt af last i kg
        "Brændstoftype"      # diesel, benzin, el, hybrid
    ]
    
    all_columns = mandatory_columns + optional_columns
    
    # 2. Opret header række
    for col_idx, column_name in enumerate(all_columns, 1):
        cell = data_sheet.cell(row=1, column=col_idx)
        cell.value = column_name
        
        # Styling for obligatoriske felter
        if column_name in mandatory_columns:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
        else:
            cell.font = Font(bold=True, color="000000")
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"), 
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
    
    # 3. Tilføj eksempel data (række 2-4 med forskellige slutadresse formater)
    example_rows = [
        [
            "Nørregade 10",      # Adresse
            "1000",              # Postnummer
            "København",         # PostDistrikt
            "Rugvænget 18, 8444 Grenå",  # SlutAdresse (almindelig adresse)
            "ABC Transport",     # Navn
            "2024-08-26",        # Dato
            "Lastbil",          # KøretøjsType
            "2500",             # LastVægt
            "diesel"            # Brændstoftype
        ],
        [
            "Vesterbrogade 5",   # Adresse
            "1620",              # Postnummer
            "København V",       # PostDistrikt
            "56.4167,10.7833",   # SlutAdresse (koordinater)
            "XYZ Logistik",      # Navn
            "2024-08-27",        # Dato
            "Varebil",          # KøretøjsType
            "1200",             # LastVægt
            "diesel"            # Brændstoftype
        ],
        [
            "Søndergade 20",     # Adresse
            "8000",              # Postnummer
            "Aarhus C",          # PostDistrikt
            "1061",              # SlutAdresse (anlæg ID)
            "DEF Kørsel",        # Navn
            "2024-08-28",        # Dato
            "Lastbil",          # KøretøjsType
            "3000",             # LastVægt
            "el"                # Brændstoftype
        ]
    ]
    
    # Tilføj eksempel rækker
    for row_idx, example_data in enumerate(example_rows, 2):
        for col_idx, value in enumerate(example_data, 1):
            cell = data_sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"), 
                bottom=Side(border_style="thin")
            )
    
    # 4. Data validering
    # Brændstoftype validering
    fuel_validation = DataValidation(
        type="list",
        formula1='"diesel,benzin,el,hybrid"',
        allow_blank=True
    )
    fuel_validation.error = "Vælg venligst en gyldig brændstoftype"
    fuel_validation.errorTitle = "Ugyldig brændstoftype"
    data_sheet.add_data_validation(fuel_validation)
    fuel_validation.add(f"I3:I1000")  # Brændstoftype kolonne
    
    # KøretøjsType validering
    vehicle_validation = DataValidation(
        type="list", 
        formula1='"Personbil,Lastbil,Varebil,Trailer"',
        allow_blank=True
    )
    vehicle_validation.error = "Vælg venligst en gyldig køretøjstype"
    vehicle_validation.errorTitle = "Ugyldig køretøjstype"
    data_sheet.add_data_validation(vehicle_validation)
    vehicle_validation.add(f"G3:G1000")  # KøretøjsType kolonne
    
    # Postnummer validering (4 cifre)
    postal_validation = DataValidation(
        type="whole",
        operator="between",
        formula1=1000,
        formula2=9999,
        allow_blank=False
    )
    postal_validation.error = "Postnummer skal være mellem 1000 og 9999"
    postal_validation.errorTitle = "Ugyldigt postnummer"
    data_sheet.add_data_validation(postal_validation)
    postal_validation.add(f"B3:B1000")  # Postnummer kolonne
    
    # 5. Kolonne bredder
    column_widths = [
        20,  # Adresse
        12,  # Postnummer
        15,  # PostDistrikt
        25,  # SlutAdresse (bredere for fleksible formater)
        20,  # Navn
        12,  # Dato
        15,  # KøretøjsType
        12,  # LastVægt
        15   # Brændstoftype
    ]
    
    for col_idx, width in enumerate(column_widths, 1):
        data_sheet.column_dimensions[data_sheet.cell(row=1, column=col_idx).column_letter].width = width
    
    # 6. Beskyt header række
    data_sheet.protection.sheet = True
    data_sheet.protection.password = "jordtransport2024"
    
    # Tillad redigering af data celler (række 2 og frem)
    for row in range(2, 1001):  # Tillad op til 1000 rækker
        for col in range(1, len(all_columns) + 1):
            data_sheet.cell(row=row, column=col).protection = None
    
    # 7. Opret Instruktioner ark
    instructions_sheet = wb.create_sheet("Instruktioner", 1)
    
    # Instruktions indhold
    instructions_content = [
        ["JORD TRANSPORT - EXCEL SKABELON VEJLEDNING", ""],
        ["", ""],
        ["OBLIGATORISKE FELTER (røde kolonner):", ""],
        ["", ""],
        ["Adresse", "Startadressens gade og husnummer"],
        ["Postnummer", "Startadressens postnummer (4 cifre)"],
        ["PostDistrikt", "Startadressens by/distrikt"],
        ["SlutAdresse", "Slutadresse i et af tre formater (se nedenfor)"],
        ["", ""],
        ["VALGFRIE FELTER (grå kolonner):", ""],
        ["", ""],
        ["Navn", "Virksomheds- eller projektnavn"],
        ["Dato", "Transport dato (format: YYYY-MM-DD)"],
        ["KøretøjsType", "Type køretøj: Personbil, Lastbil, Varebil, Trailer"],
        ["LastVægt", "Vægt af last i kg (kun tal)"],
        ["Brændstoftype", "diesel, benzin, el eller hybrid"],
        ["", ""],
        ["SLUTADRESSE FORMATER:", ""],
        ["", ""],
        ["Format 1: Almindelig adresse", "Eksempel: 'Rugvænget 18, 8444 Grenå'"],
        ["Format 2: Koordinater", "Eksempel: '56.4167,10.7833' (breddegrad,længdegrad)"],
        ["Format 3: Anlæg ID", "Eksempel: '1061' (slås op i adressedatabase)"],
        ["", ""],
        ["GYLDIGE ANLÆG ID'er (Format 3):", ""],
        ["", ""],
        ["1061", "Gert Svith, Birkesig Grusgrav - Rugvænget 18, 8444 Grenå"],
        ["1013", "JJ Grus A/S (Kalbygård Grusgrav) - Hovedvejen 24A, 8670 Låsby"],
        ["1327", "Johs. Sørensen & Sønner A/S - Holmstrupgårdvej 9, 8220 Brabrand"],
        ["2191", "JJ Grus A/S (Ans) - Søndermarksgade 43, 8643 Ans"],
        ["1901", "EHJ Energi & Miljø A/S - Hadstenvej 16, 8940 Randers SV"],
        ["", ""],
        ["EKSEMPEL PÅ KORREKT UDFYLDTE RÆKKER:", ""],
        ["", ""],
        ["Eksempel 1 (almindelig adresse):", ""],
        ["Adresse: Nørregade 10", "Postnummer: 1000"],
        ["PostDistrikt: København", "SlutAdresse: Rugvænget 18, 8444 Grenå"],
        ["", ""],
        ["Eksempel 2 (koordinater):", ""],
        ["Adresse: Vesterbrogade 5", "Postnummer: 1620"],
        ["PostDistrikt: København V", "SlutAdresse: 56.4167,10.7833"],
        ["", ""],
        ["Eksempel 3 (anlæg ID):", ""],
        ["Adresse: Søndergade 20", "Postnummer: 8000"],
        ["PostDistrikt: Aarhus C", "SlutAdresse: 1061"],
        ["", ""],
        ["BEMÆRKNINGER:", ""],
        ["", ""],
        ["• Header-rækken er beskyttet mod ændringer", ""],
        ["• Obligatoriske felter SKAL udfyldes for hver række", ""],
        ["• SlutAdresse kan være adresse, koordinater eller anlæg ID", ""],
        ["• Postnummer skal være mellem 1000 og 9999", ""],
        ["• Data validering er sat op for visse felter", ""],
        ["• Gem filen som .xlsx eller .csv format for upload", ""],
        ["• Maksimal filstørrelse: 10MB", ""],
    ]
    
    # Tilføj instruktioner til ark
    for row_idx, (col1, col2) in enumerate(instructions_content, 1):
        cell_a = instructions_sheet.cell(row=row_idx, column=1)
        cell_b = instructions_sheet.cell(row=row_idx, column=2)
        
        cell_a.value = col1
        cell_b.value = col2
        
        # Styling for headings
        if "VEJLEDNING" in str(col1):
            cell_a.font = Font(bold=True, size=16, color="FFFFFF")
            cell_a.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        elif str(col1).endswith(":") and col1 != "":
            cell_a.font = Font(bold=True, size=12, color="1976D2")
        elif col1.isdigit():
            cell_a.font = Font(bold=True, color="D32F2F")
            cell_b.font = Font(color="333333")
    
    # Opsæt kolonne bredder for instruktioner
    instructions_sheet.column_dimensions['A'].width = 25
    instructions_sheet.column_dimensions['B'].width = 50
    
    # 8. Opret Validerings ark med ModtageranlægID reference
    validation_sheet = wb.create_sheet("Gyldige_IDs", 2)
    
    # Data for validering
    validation_data = [
        ["ModtageranlægID", "Navn", "Adresse"],
        [1061, "Gert Svith, Birkesig Grusgrav", "Rugvænget 18, 8444 Grenå"],
        [1013, "JJ Grus A/S (Kalbygård Grusgrav)", "Hovedvejen 24A, 8670 Låsby"],
        [1327, "Johs. Sørensen & Sønner A/S, Ren depotjord", "Holmstrupgårdvej 9, 8220 Brabrand"],
        [2191, "JJ Grus A/S (Ans)", "Søndermarksgade 43, 8643 Ans"],
        [1901, "EHJ Energi & Miljø A/S - Let forurenet jord", "Hadstenvej 16, 8940 Randers SV"],
    ]
    
    for row_idx, row_data in enumerate(validation_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = validation_sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            if row_idx == 1:  # Header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            
            cell.border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
    
    # Opsæt kolonne bredder for validering
    validation_sheet.column_dimensions['A'].width = 18
    validation_sheet.column_dimensions['B'].width = 35
    validation_sheet.column_dimensions['C'].width = 30
    
    # Gem template
    template_path = Path(__file__).parent / "jord_transport_template.xlsx"
    wb.save(str(template_path))
    
    print(f"Excel skabelon oprettet: {template_path}")
    print("Header-beskyttelse: Aktiveret med password 'jordtransport2024'")
    print("Data validering: Sat op for brændstoftype, køretøjstype og postnummer")
    print("Arkene: Data, Instruktioner, Gyldige_IDs")
    
    return str(template_path)

def create_csv_template():
    """
    Opretter en enkel CSV skabelon som alternativ til Excel.
    """
    
    # Definer kolonner
    mandatory_columns = ["Adresse", "Postnummer", "PostDistrikt", "SlutAdresse"]
    optional_columns = ["Navn", "Dato", "KøretøjsType", "LastVægt", "Brændstoftype"]
    all_columns = mandatory_columns + optional_columns
    
    # Opret DataFrame med header og eksempler (alle tre formater)
    example_data = {
        "Adresse": ["Nørregade 10", "Vesterbrogade 5", "Søndergade 20"],
        "Postnummer": [1000, 1620, 8000], 
        "PostDistrikt": ["København", "København V", "Aarhus C"],
        "SlutAdresse": ["Rugvænget 18, 8444 Grenå", "56.4167,10.7833", "1061"],
        "Navn": ["ABC Transport", "XYZ Logistik", "DEF Kørsel"],
        "Dato": ["2024-08-26", "2024-08-27", "2024-08-28"],
        "KøretøjsType": ["Lastbil", "Varebil", "Lastbil"],
        "LastVægt": [2500, 1200, 3000],
        "Brændstoftype": ["diesel", "diesel", "el"]
    }
    
    df = pd.DataFrame(example_data)
    
    # Gem CSV template
    csv_path = Path(__file__).parent / "jord_transport_template.csv"
    df.to_csv(str(csv_path), index=False, encoding='utf-8')
    
    print(f"CSV skabelon oprettet: {csv_path}")
    
    return str(csv_path)

if __name__ == "__main__":
    print("Opretter Jord Transport skabeloner...")
    print("=" * 50)
    
    # Opret templates mappe hvis den ikke eksisterer
    template_dir = Path(__file__).parent
    template_dir.mkdir(exist_ok=True)
    
    # Opret begge skabeloner
    excel_path = create_excel_template()
    csv_path = create_csv_template()
    
    print("=" * 50)
    print("Skabeloner oprettet succesfuldt!")
    print(f"Excel: {excel_path}")
    print(f"CSV: {csv_path}")